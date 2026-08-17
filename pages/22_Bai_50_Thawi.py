import streamlit as st
import sys, os, io
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import style
from utils import bai50_api

bai50_api.start(port=8504)

st.set_page_config(
    page_title="ใบ 50 ทวิ",
    page_icon="📋",
    layout="wide",
)
style.inject()
style.back_home()

st.markdown(
    '<span style="background:linear-gradient(135deg,#1a3a6b,#2c5aa0);color:white;'
    'border-radius:20px;padding:4px 14px;font-size:12px;font-weight:700;">Withholding Tax Certificate</span>',
    unsafe_allow_html=True,
)
st.title("📋 ใบ 50 ทวิ — หนังสือรับรองการหักภาษี ณ ที่จ่าย")
st.caption("กรอกข้อมูล แล้วพิมพ์หรือบันทึกเป็น PDF ได้ทันที — ออกทั้งฉบับที่ 1 และ ฉบับที่ 2 พร้อมกัน")


def create_bai50_template() -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ข้อมูลใบ50ทวิ"

    columns = [
        ("payee_name",       "ชื่อผู้รับเงิน *",                         28),
        ("payee_tax_id",     "เลขประจำตัวฯ 13 หลัก *",                   20),
        ("payee_addr",       "ที่อยู่ผู้รับเงิน",                          38),
        ("book_no",          "เล่มที่",                                    9),
        ("doc_no",           "เลขที่เอกสาร",                               14),
        ("form_type",        "ลำดับที่ในแบบ",                             17),
        ("r1_date",          "วันที่ 40(1) เงินเดือน",                     16),
        ("r1_amount",        "จำนวน 40(1)",                               13),
        ("r1_tax",           "ภาษีหัก 40(1)",                             13),
        ("r2_date",          "วันที่ 40(2) ค่าธรรมเนียม",                 16),
        ("r2_amount",        "จำนวน 40(2)",                               13),
        ("r2_tax",           "ภาษีหัก 40(2)",                             13),
        ("r3_date",          "วันที่ 40(3) ลิขสิทธิ์",                    16),
        ("r3_amount",        "จำนวน 40(3)",                               13),
        ("r3_tax",           "ภาษีหัก 40(3)",                             13),
        ("r4a_date",         "วันที่ 40(4)(ก) ดอกเบี้ย",                  16),
        ("r4a_amount",       "จำนวน 40(4)(ก)",                            13),
        ("r4a_tax",          "ภาษีหัก 40(4)(ก)",                          13),
        ("r5_date",          "วันที่ ม.3 เตรส",                           14),
        ("r5_amount",        "จำนวน ม.3 เตรส",                           13),
        ("r5_tax",           "ภาษีหัก ม.3 เตรส",                         13),
        ("r6_desc",          "อื่นๆ ระบุประเภท",                          18),
        ("r6_date",          "วันที่ อื่นๆ",                               12),
        ("r6_amount",        "จำนวน อื่นๆ",                               13),
        ("r6_tax",           "ภาษีหัก อื่นๆ",                             13),
        ("fund1",            "กบข./กสจ. (บาท)",                           15),
        ("fund2",            "ประกันสังคม (บาท)",                          15),
        ("fund3",            "กองทุนสำรองฯ (บาท)",                        17),
        ("condition",        "เงื่อนไข (1/2/3/4)",                        15),
        ("condition_other",  "เงื่อนไขอื่น (ถ้าเลือก 4)",                 20),
        ("signer_name",      "ชื่อผู้ลงนาม",                              20),
        ("issue_date",       "วันที่ออกหนังสือ",                           16),
    ]

    hdr_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", name="TH Sarabun New", size=12)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wr  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for col_i, (_, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_i)].width = width

    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"

    # Text format for tax ID column (column B = index 2) — prevent Excel from treating 13-digit IDs as numbers
    tax_id_col = next(c+1 for c, (k, _, _) in enumerate(columns) if k == "payee_tax_id")
    for row_i in range(2, 502):
        ws.cell(row=row_i, column=tax_id_col).number_format = "@"

    # Number format for amount columns
    num_cols = {c+1 for c, (k, _, _) in enumerate(columns) if "amount" in k or k.endswith("_tax") or k in ("fund1","fund2","fund3")}
    for col_i in num_cols:
        for row_i in range(2, 502):
            ws.cell(row=row_i, column=col_i).number_format = "#,##0.00"

    # Dropdown: form_type (column 6 = F)
    dv_ft = DataValidation(
        type="list",
        formula1='"ภ.ง.ด.1ก,ภ.ง.ด.1ก พิเศษ,ภ.ง.ด.2,ภ.ง.ด.3,ภ.ง.ด.2ก,ภ.ง.ด.3ก,ภ.ง.ด.53"',
        allow_blank=True,
        showErrorMessage=False,
    )
    dv_ft.sqref = "F2:F501"
    ws.add_data_validation(dv_ft)

    # Dropdown: condition (column 29 = AC)
    dv_cond = DataValidation(
        type="list",
        formula1='"1,2,3,4"',
        allow_blank=True,
        showErrorMessage=False,
    )
    dv_cond.sqref = "AC2:AC501"
    ws.add_data_validation(dv_cond)

    # ===== Instructions sheet =====
    ws2 = wb.create_sheet("คำอธิบาย")
    rows_info = [
        ["คอลัมน์",             "คำอธิบาย",                                          "ตัวอย่าง",                                    "หมายเหตุ"],
        ["payee_name",          "ชื่อผู้รับเงิน (ผู้ถูกหัก)",                          "นาย สมชาย ใจดี",                              "จำเป็น *"],
        ["payee_tax_id",        "เลขประจำตัวผู้เสียภาษีอากร 13 หลัก",                 "1234567890123",                               "จำเป็น * ไม่ต้องใส่ขีด"],
        ["payee_addr",          "ที่อยู่ผู้รับเงิน",                                    "123 ถ.สุขุมวิท แขวงคลองตัน เขตคลองเตย กทม.", ""],
        ["book_no",             "เล่มที่ (ไม่บังคับ)",                                  "",                                            ""],
        ["doc_no",              "เลขที่เอกสาร",                                        "001/2568",                                    ""],
        ["form_type",           "ลำดับที่ในแบบ",                                        "ภ.ง.ด.1ก",                                   "เลือกจาก dropdown หรือพิมพ์"],
        ["r1_date–r1_tax",      "มาตรา 40(1) เงินเดือน ค่าจ้าง ฯลฯ",                  "2568 / 50000 / 2500",                         "วันที่เป็นปี พ.ศ. หรือ วว/ดด/ปปปป"],
        ["r2_date–r2_tax",      "มาตรา 40(2) ค่าธรรมเนียม ค่านายหน้า ฯลฯ",            "",                                            ""],
        ["r3_date–r3_tax",      "มาตรา 40(3) ค่าแห่งลิขสิทธิ์ ฯลฯ",                  "",                                            ""],
        ["r4a_date–r4a_tax",    "มาตรา 40(4)(ก) ดอกเบี้ย ฯลฯ",                        "",                                            ""],
        ["r5_date–r5_tax",      "มาตรา 3 เตรส (ค่าเช่า ค่าบริการ ค่าจ้างทำของ ฯลฯ)",  "",                                            ""],
        ["r6_desc–r6_tax",      "อื่น ๆ ระบุประเภท + วันที่ + จำนวน + ภาษีหัก",       "",                                            ""],
        ["fund1",               "กบข./กสจ./กองทุนสงเคราะห์ครูฯ (บาท)",                "",                                            "ถ้าไม่มีให้เว้นว่าง"],
        ["fund2",               "กองทุนประกันสังคม (บาท)",                              "",                                            ""],
        ["fund3",               "กองทุนสำรองเลี้ยงชีพ (บาท)",                          "",                                            ""],
        ["condition",           "เงื่อนไขการออกหนังสือ",                               "1",                                           "1=หักณที่จ่าย 2=ออกให้ตลอดไป 3=ออกให้ครั้งเดียว 4=อื่นๆ"],
        ["condition_other",     "เงื่อนไขอื่นๆ (ใช้เมื่อ condition=4)",                "",                                            ""],
        ["signer_name",         "ชื่อผู้ลงนาม",                                        "นาย สมศักดิ์ มั่นคง",                         "ถ้าว่างจะใช้ค่าจากฟอร์มหลัก"],
        ["issue_date",          "วันที่ออกหนังสือรับรองฯ",                              "27/05/2568",                                  "ถ้าว่างจะใช้ค่าจากฟอร์มหลัก"],
        [],
        ["หมายเหตุ:", "ข้อมูลผู้จ่ายเงิน (บริษัท/ผู้มีหน้าที่หัก) ให้เลือกจากโปรไฟล์ในแท็บ 'กรอกข้อมูล' ก่อน — ข้อมูลจะถูกใช้สำหรับทุกใบในชุดนั้น", "", ""],
    ]

    i2_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    i2_font = Font(bold=True, color="FFFFFF", name="TH Sarabun New", size=12)
    for row_data in rows_info:
        ws2.append(row_data)
    for cell in ws2[1]:
        cell.font = i2_font
        cell.fill = i2_fill

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 38
    ws2.column_dimensions["D"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


col_title, col_dl = st.columns([5, 2])
with col_dl:
    tpl_bytes = create_bai50_template()
    if tpl_bytes:
        st.download_button(
            label="📥 ดาวน์โหลด Template Excel (Batch)",
            data=tpl_bytes,
            file_name="bai50_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("ติดตั้ง openpyxl เพื่อใช้งาน template: `pip install openpyxl`")

html_path = Path(__file__).parent.parent / "tools" / "bai_50_thawi.html"
html_content = html_path.read_text(encoding="utf-8")

st.components.v1.html(html_content, height=1900, scrolling=True)
